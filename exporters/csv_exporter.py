import logging
import re
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import settings
from models.lead import Lead
from utils.file_utils import ensure_dir

logger = logging.getLogger(__name__)

SHEET_NAME = "ScrapBro Leads"

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
ODD_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DATA_FONT = Font(color="000000", size=10)
TOTALS_FONT = Font(color="FFFFFF", bold=True, size=10)
EDGE_SIDE = Side(style="thin", color="CCCCCC")

# A column is (header, width, getter) where getter(lead) -> cell value.
Column = tuple[str, int, Callable[[Lead], object]]


def _field(name: str) -> Callable[[Lead], object]:
    """Getter that reads a base Lead field via to_dict()."""
    return lambda lead: lead.to_dict().get(name, "")


def _raw(key: str) -> Callable[[Lead], object]:
    """Getter that reads a key from a Lead's raw_data."""
    return lambda lead: (lead.raw_data or {}).get(key, "")


def _entity_type(lead: Lead) -> str:
    """Human-readable entity type from raw_data (Física / Jurídica)."""
    value = (lead.raw_data or {}).get("entity_type", "")
    return {"fisica": "Física", "juridica": "Jurídica"}.get(value, "")


BASE_COLUMNS: list[Column] = [
    ("Name", 30, _field("name")),
    ("Email", 30, _field("email")),
    ("Phone", 18, _field("phone")),
    ("Website", 35, _field("website")),
    ("Address", 40, _field("address")),
    ("Category", 20, _field("category")),
    ("Rating", 10, _field("rating")),
    ("Source", 15, _field("source")),
]

DATEAS_COLUMNS: list[Column] = [
    ("DNI", 15, _raw("dni")),
    ("CUIT/CUIL", 18, _raw("cuit")),
    ("Edad", 8, _raw("age")),
    ("Provincia", 20, _raw("province")),
    ("Localidad", 25, _raw("locality")),
    ("Tipo", 12, _entity_type),
]

# Appended only when a multi-query run tags leads with their originating query.
QUERY_COLUMN: Column = ("Query", 25, _raw("query"))

HEADER_ROW_HEIGHT = 20
DATA_ROW_HEIGHT = 16


def build_output_filename(query: str, output_format: str = "csv") -> str:
    """Build the output filename from the user's query, with the right extension.

    Args:
        query: The original query (e.g. "inmobiliarias lujan", or a multi-query
            like "inmobiliaria lujan -- santiago gomez").
        output_format: Selected output format, ``"csv"`` (Excel) or ``"json"``.
            Drives the extension so JSON output never gets an ``.xlsx`` name.

    Returns:
        A safe filename with the correct extension, e.g. "inmobiliarias_lujan.xlsx"
        for ``"csv"`` or "inmobiliarias_lujan.json" for ``"json"``.
    """
    def _slug(part: str) -> str:
        cleaned = re.sub(r"[^\w\s-]", "", part.lower())
        cleaned = re.sub(r"\s+", "_", cleaned.strip())
        return re.sub(r"_+", "_", cleaned).strip("_")

    parts = [_slug(p) for p in query.split("--")]
    name = "__".join(p for p in parts if p)
    extension = "json" if output_format == "json" else "xlsx"
    return f"{name[:80] or 'scrapbro'}.{extension}"


class CSVExporter:
    """Exports leads to a styled Excel (.xlsx) workbook.

    Kept under the historical CSVExporter name: the CLI's --output csv now
    produces .xlsx because a formatted sheet is more useful than plain CSV.
    Columns are dynamic: the six Dateas-specific columns (DNI, CUIT, age,
    province, locality, entity type) are appended only when the result set
    contains leads from the ``dateas`` source.
    """

    def export(self, leads: list[Lead], filename: str) -> str | None:
        """Save leads to a styled .xlsx file in the configured output directory.

        Args:
            leads: List of Lead objects to export.
            filename: Base filename; a .csv suffix is swapped to .xlsx.

        Returns:
            Absolute path to the generated .xlsx file, or None when there are no
            leads (no empty workbook is written).
        """
        if not leads:
            logger.warning("No leads to export — skipping file creation")
            return None

        path = ensure_dir(Path(settings.OUTPUT_DIR)) / Path(filename).with_suffix(".xlsx").name

        columns = self._columns_for(leads)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

        self._write_header(sheet, columns)
        for row_index, lead in enumerate(leads, start=2):
            self._write_data_row(sheet, row_index, lead, columns)
        totals_row = len(leads) + 2
        self._write_totals_row(sheet, totals_row, leads, columns)
        self._apply_outer_border(sheet, totals_row, len(columns))

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

        workbook.save(path)
        logger.info("XLSX exported: %s (%d leads, %d columns)", path, len(leads), len(columns))
        return str(path.resolve())

    @staticmethod
    def _columns_for(leads: list[Lead]) -> list[Column]:
        """Return the column set for this result.

        Base columns are always present. The six Dateas columns are appended
        when any lead is from the ``dateas`` source; a ``Query`` column is
        appended when leads carry an originating query (multi-query runs).
        """
        columns = list(BASE_COLUMNS)
        if any(lead.source == "dateas" for lead in leads):
            columns += DATEAS_COLUMNS
        if any((lead.raw_data or {}).get("query") for lead in leads):
            columns.append(QUERY_COLUMN)
        return columns

    @staticmethod
    def _write_header(sheet, columns: list[Column]) -> None:
        """Write the styled header row (dark blue, white bold text)."""
        for col_index, (header, width, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            sheet.column_dimensions[get_column_letter(col_index)].width = width
        sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT

    @staticmethod
    def _write_data_row(sheet, row_index: int, lead: Lead, columns: list[Column]) -> None:
        """Write one lead row with alternating background fill."""
        for col_index, (_, _, getter) in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=getter(lead))
            cell.font = DATA_FONT
            if row_index % 2 != 0:
                cell.fill = ODD_ROW_FILL
        sheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT

    @staticmethod
    def _write_totals_row(sheet, row_index: int, leads: list[Lead], columns: list[Column]) -> None:
        """Write the merged totals row at the end of the data."""
        total = len(leads)
        last_col = len(columns)

        def pct(part: int) -> str:
            return f"{round(part * 100 / total)}%" if total else "0%"

        with_email = sum(1 for lead in leads if lead.email)
        with_phone = sum(1 for lead in leads if lead.phone)

        if any(header == "DNI" for header, _, _ in columns):
            with_dni = sum(1 for lead in leads if (lead.raw_data or {}).get("dni"))
            with_cuit = sum(1 for lead in leads if (lead.raw_data or {}).get("cuit"))
            summary = (
                f"Total: {total} leads  |  Con DNI: {with_dni} ({pct(with_dni)})  |  "
                f"Con CUIT: {with_cuit} ({pct(with_cuit)})  |  "
                f"Con email: {with_email} ({pct(with_email)})  |  "
                f"Con teléfono: {with_phone} ({pct(with_phone)})"
            )
        else:
            with_web = sum(1 for lead in leads if lead.website)
            summary = (
                f"Total: {total} leads  |  Con email: {with_email} ({pct(with_email)})  |  "
                f"Con teléfono: {with_phone} ({pct(with_phone)})  |  Con web: {with_web} ({pct(with_web)})"
            )

        sheet.merge_cells(
            start_row=row_index, start_column=1, end_row=row_index, end_column=last_col
        )
        cell = sheet.cell(row=row_index, column=1, value=summary)
        cell.fill = HEADER_FILL
        cell.font = TOTALS_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_index in range(2, last_col + 1):
            sheet.cell(row=row_index, column=col_index).fill = HEADER_FILL

    @staticmethod
    def _apply_outer_border(sheet, last_row: int, last_col: int) -> None:
        """Draw a solid grey border around the full used range, no inner borders."""
        for row in range(1, last_row + 1):
            for col in (1, last_col):
                cell = sheet.cell(row=row, column=col)
                cell.border = Border(
                    left=EDGE_SIDE if col == 1 else None,
                    right=EDGE_SIDE if col == last_col else None,
                    top=EDGE_SIDE if row == 1 else None,
                    bottom=EDGE_SIDE if row == last_row else None,
                )
        for col in range(2, last_col):
            sheet.cell(row=1, column=col).border = Border(top=EDGE_SIDE)
            bottom = sheet.cell(row=last_row, column=col)
            bottom.border = Border(bottom=EDGE_SIDE)
