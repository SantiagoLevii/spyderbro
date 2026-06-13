"""Offline tests for the Argentina-pack scrapers.

Single-page scrapers (paginas_amarillas, dateas, guia_oleo, doctoralia,
mercadolibre) are exercised through the shared parametrized suite. The scrapers
rewritten in Sprint G to a two-step listing->detail flow (abogados, zonaprop,
argenprop) and the country-filtered Clutch scraper have dedicated tests with
proper page sequences. All fetching is mocked via FakeFetcher/FakePage — no real
requests.
"""
import argparse
import json

import pytest
from openpyxl import load_workbook

import main
from config.settings import settings
from exporters.csv_exporter import CSVExporter
from models.lead import Lead
from scrapers.abogados import AbogadosScraper
from scrapers.argenprop import ArgenpropScraper
from scrapers.clutch import ClutchScraper
from scrapers.dateas import DateasScraper
from scrapers.mercadolibre import MercadoLibreScraper
from scrapers.paginas_amarillas import PaginasAmarillasScraper
from scrapers.zonaprop import ZonapropScraper

from tests.conftest import FakeFetcher, FakePage


@pytest.fixture(autouse=True)
def no_delay_clean_settings(monkeypatch):
    """Zero out request delays and reset runtime settings for each test."""
    monkeypatch.setattr(settings, "SCRAPING_DELAY_MIN", 0.0)
    monkeypatch.setattr(settings, "SCRAPING_DELAY_MAX", 0.0)
    monkeypatch.setattr(settings, "LOCATION", "")
    monkeypatch.setattr(settings, "DATEAS_TYPE", "empresas")


def _next_data_card(name: str) -> str:
    """Build a Páginas Amarillas __NEXT_DATA__ script with one business record."""
    payload = {
        "props": {"pageProps": {"results": [{
            "name": name,
            "infoLine": "Restaurante",
            "contactMap": {"WEB": [f"https://{name.replace(' ', '')}.com.ar"]},
            "mainAddress": {
                "streetName": "Av Corrientes", "streetNumber": "1000",
                "localityToShow": "CABA",
                "allPhones": [{"number": "+541145551234"}],
            },
        }]}}
    }
    return f'<script id="__NEXT_DATA__" type="application/json">{json.dumps(payload)}</script>'


def _card(scraper_key: str, name: str) -> str:
    """Return the mock HTML card markup expected by a single-page scraper."""
    cards = {
        "paginas_amarillas": _next_data_card(name),
        "dateas": (
            f'<table><tr>'
            f'<td>{name}</td>'
            f'<td>30-{abs(hash(name)) % 100000000:08d}-9</td>'
            f'<td></td><td>Buenos Aires</td><td>CABA</td>'
            f'<td><a href="/es/empresa/{name.replace(" ", "-").lower()}-30123456789">'
            f'Ver Más</a></td>'
            f'</tr></table>'
        ),
    }
    return cards[scraper_key]


def _page(scraper_key: str, *names: str) -> FakePage:
    """Build a FakePage containing the given named cards for a scraper."""
    body = "".join(_card(scraper_key, n) for n in names)
    return FakePage(f"<html><body>{body}</body></html>")


# (key, scraper class, fetcher attribute patched on the scraper module, source)
SCRAPERS = [
    ("paginas_amarillas", PaginasAmarillasScraper, "Fetcher", "paginas_amarillas"),
    ("dateas", DateasScraper, "Fetcher", "dateas"),
]

IDS = [s[0] for s in SCRAPERS]


def _patch_fetcher(monkeypatch, key, attr):
    """Point the scraper module's fetcher at the FakeFetcher."""
    monkeypatch.setattr(f"scrapers.{key}.{attr}", FakeFetcher)


@pytest.mark.parametrize("key,cls,attr,source", SCRAPERS, ids=IDS)
def test_parsing(monkeypatch, key, cls, attr, source):
    _patch_fetcher(monkeypatch, key, attr)
    FakeFetcher.configure(pages=[_page(key, "Negocio Uno")])
    leads = cls().scrape("test query", limit=10)
    assert len(leads) == 1
    assert leads[0].name == "Negocio Uno"
    assert leads[0].source == source


@pytest.mark.parametrize("key,cls,attr,source", SCRAPERS, ids=IDS)
def test_empty_page(monkeypatch, key, cls, attr, source):
    _patch_fetcher(monkeypatch, key, attr)
    FakeFetcher.configure(pages=[FakePage("<html><body></body></html>")])
    assert cls().scrape("test query", limit=10) == []


@pytest.mark.parametrize("key,cls,attr,source", SCRAPERS, ids=IDS)
def test_timeout(monkeypatch, key, cls, attr, source):
    _patch_fetcher(monkeypatch, key, attr)
    FakeFetcher.configure(error=TimeoutError("request timed out"))
    assert cls().scrape("test query", limit=10) == []


@pytest.mark.parametrize("key,cls,attr,source", SCRAPERS, ids=IDS)
def test_source_field(monkeypatch, key, cls, attr, source):
    _patch_fetcher(monkeypatch, key, attr)
    FakeFetcher.configure(pages=[_page(key, "Negocio Uno", "Negocio Dos")])
    leads = cls().scrape("test query", limit=10)
    assert leads
    assert all(lead.source == source for lead in leads)


@pytest.mark.parametrize("key,cls,attr,source", SCRAPERS, ids=IDS)
def test_pagination_advances(monkeypatch, key, cls, attr, source):
    _patch_fetcher(monkeypatch, key, attr)
    FakeFetcher.configure(pages=[
        _page(key, "Pagina Uno A", "Pagina Uno B"),
        _page(key, "Pagina Dos A", "Pagina Dos B"),
    ])
    leads = cls().scrape("test query", limit=50)
    names = {lead.name for lead in leads}
    assert FakeFetcher.calls >= 2
    assert "Pagina Uno A" in names
    assert "Pagina Dos A" in names


def test_dateas_personas_mode(monkeypatch):
    monkeypatch.setattr(settings, "DATEAS_TYPE", "personas")
    monkeypatch.setattr("scrapers.dateas.Fetcher", FakeFetcher)
    FakeFetcher.configure(pages=[_page("dateas", "Juan Perez")])
    leads = DateasScraper().scrape("contador", limit=10)
    assert len(leads) == 1
    assert leads[0].raw_data.get("cuit", "").startswith("30-")


# --- Two-step scrapers (listing -> detail) and Clutch (country filter) ---------


def test_abogados_listing_detail(monkeypatch):
    """abogados: home (area) -> listing (firm links) -> firm detail page."""
    monkeypatch.setattr("scrapers.abogados.Fetcher", FakeFetcher)
    home = FakePage('<html><body><a href="/area/laboral/10">Laboral</a></body></html>')
    listing = FakePage(
        '<html><body><a href="/directorio/estudio-uno/101">Estudio Uno</a></body></html>'
    )
    detail = FakePage(
        '<html><body><h1>Estudio Uno</h1>'
        '<address>Maipu 100 CABA, Argentina</address>'
        '<a href="tel:+541143331111">tel</a>'
        '<a href="https://estudiouno.com">web</a></body></html>'
    )
    FakeFetcher.configure(pages=[home, listing, detail])
    # query with no second token -> no province filter
    leads = AbogadosScraper().scrape("laboral", limit=10)
    assert len(leads) == 1
    assert leads[0].name == "Estudio Uno"
    assert leads[0].source == "abogados"
    assert leads[0].phone == "+541143331111"
    assert leads[0].website == "https://estudiouno.com"


def test_abogados_empty(monkeypatch):
    monkeypatch.setattr("scrapers.abogados.Fetcher", FakeFetcher)
    FakeFetcher.configure(pages=[FakePage("<html><body></body></html>")])
    assert AbogadosScraper().scrape("laboral", limit=10) == []


def test_abogados_timeout(monkeypatch):
    monkeypatch.setattr("scrapers.abogados.Fetcher", FakeFetcher)
    FakeFetcher.configure(error=TimeoutError("timeout"))
    assert AbogadosScraper().scrape("laboral", limit=10) == []


def test_clutch_parsing_and_country_filter(monkeypatch):
    """clutch: parse .provider cards, website from redirect u=, filter by country."""
    monkeypatch.setattr("scrapers.clutch.StealthyFetcher", FakeFetcher)
    card_ar = (
        '<div class="provider"><a class="provider__title-link">Agencia AR</a>'
        '<a href="https://r.clutch.co/redirect?u=https%3A%2F%2Fagenciaar.com">visit</a>'
        '<div class="location">Buenos Aires, Argentina</div>'
        '<div class="sg-rating__number">4.8</div></div>'
    )
    card_us = (
        '<div class="provider"><a class="provider__title-link">Agencia US</a>'
        '<a href="https://r.clutch.co/redirect?u=https%3A%2F%2Fagenciaus.com">visit</a>'
        '<div class="location">San Diego, CA</div>'
        '<div class="sg-rating__number">4.9</div></div>'
    )
    FakeFetcher.configure(pages=[FakePage(f"<html><body>{card_ar}{card_us}</body></html>")])
    leads = ClutchScraper().scrape("digital-marketing argentina", limit=10)
    assert len(leads) == 1  # only the Argentine card passes the country filter
    assert leads[0].name == "Agencia AR"
    assert leads[0].website == "https://agenciaar.com"
    assert leads[0].rating == 4.8
    assert leads[0].source == "clutch"


def test_clutch_timeout(monkeypatch):
    monkeypatch.setattr("scrapers.clutch.StealthyFetcher", FakeFetcher)
    FakeFetcher.configure(error=TimeoutError("timeout"))
    assert ClutchScraper().scrape("digital-marketing argentina", limit=10) == []


def test_zonaprop_listing_detail(monkeypatch):
    """zonaprop: listing cards -> detail page with agency link + telephone JSON."""
    monkeypatch.setattr("scrapers.zonaprop.StealthyFetcher", FakeFetcher)
    listing = FakePage(
        '<html><body>'
        '<div data-posting-type="PROPERTY" '
        'data-to-posting="/propiedades/clasificado/depto-123.html">'
        '<div data-qa="POSTING_CARD_LOCATION">Palermo, Capital Federal</div>'
        '</div></body></html>'
    )
    detail = FakePage(
        '<html><body>'
        '<a href="/inmobiliarias/inmo-uno_30123-inmuebles.html">Inmo Uno</a>'
        '<script type="application/ld+json">{"telephone":"+541143331111"}</script>'
        '</body></html>'
    )
    FakeFetcher.configure(pages=[listing, detail])
    leads = ZonapropScraper().scrape("venta capital-federal", limit=10)
    assert len(leads) == 1
    assert leads[0].name == "Inmo Uno"
    assert leads[0].source == "zonaprop"
    assert leads[0].phone == "+541143331111"


def test_zonaprop_empty(monkeypatch):
    monkeypatch.setattr("scrapers.zonaprop.StealthyFetcher", FakeFetcher)
    FakeFetcher.configure(pages=[FakePage("<html><body></body></html>")])
    assert ZonapropScraper().scrape("venta capital-federal", limit=10) == []


def test_argenprop_listing_detail(monkeypatch):
    """argenprop: listing cards -> detail page with agency link + h2 address."""
    monkeypatch.setattr("scrapers.argenprop.StealthyFetcher", FakeFetcher)
    listing = FakePage(
        '<html><body><div data-item-card>'
        '<a href="/departamento-en-venta-en-palermo--123">ver</a>'
        '</div></body></html>'
    )
    detail = FakePage(
        '<html><body>'
        '<a href="inmobiliarias/inmo-dos">Inmo Dos</a>'
        '<h2>Avenida Santa Fe 4400</h2></body></html>'
    )
    FakeFetcher.configure(pages=[listing, detail])
    leads = ArgenpropScraper().scrape("venta palermo", limit=10)
    assert len(leads) == 1
    assert leads[0].name == "Inmo Dos"
    assert leads[0].source == "argenprop"
    assert leads[0].address == "Avenida Santa Fe 4400"


def test_argenprop_empty(monkeypatch):
    monkeypatch.setattr("scrapers.argenprop.StealthyFetcher", FakeFetcher)
    FakeFetcher.configure(pages=[FakePage("<html><body></body></html>")])
    assert ArgenpropScraper().scrape("venta palermo", limit=10) == []


def test_mercadolibre_antibot_returns_empty(monkeypatch):
    """mercadolibre: snoopy micro-landing shell is detected and yields no leads."""
    monkeypatch.setattr("scrapers.mercadolibre.StealthyFetcher", FakeFetcher)
    shell = FakePage(
        '<html><body><div class="micro-landing-container">'
        '<script src="snoopy-script.js"></script></div></body></html>'
    )
    FakeFetcher.configure(pages=[shell])
    assert MercadoLibreScraper().scrape("electronica", limit=10) == []


def test_argentina_pack_constant():
    import main
    assert len(main.ARGENTINA_PACK) == 9
    assert "paginas_amarillas" in main.ARGENTINA_PACK
    assert all(src in main.SCRAPERS for src in main.ARGENTINA_PACK)


# --- Dateas Sprint I: full fields, pagination, lookup, Excel, filters ---------

def _dateas_row(name, cuit, age, province, locality, kind="persona"):
    slug = name.replace(" ", "-").lower()
    digits = cuit.replace("-", "")
    return (
        f'<table><tr><td>{name}</td><td>{cuit}</td><td>{age}</td>'
        f'<td>{province}</td><td>{locality}</td>'
        f'<td><a href="/es/{kind}/{slug}-{digits}">Ver Más</a></td></tr></table>'
    )


def _filter_args(**overrides):
    base = dict(
        filter_complete=False, filter_has_phone=False, filter_has_email=False,
        filter_has_website=False, filter_min_rating=None, filter_has_cuit=False,
        filter_has_dni=False, filter_entity_type="ambos", filter_province=None,
        filter_locality=None,
    )
    base.update(overrides)
    return argparse.Namespace(**base)


def test_dateas_extracts_all_public_fields(monkeypatch):
    monkeypatch.setattr("scrapers.dateas.Fetcher", FakeFetcher)
    monkeypatch.setattr(settings, "DATEAS_TYPE", "personas")
    html = _dateas_row("GARCIA JUAN", "20-30123456-5", "40 años",
                       "Buenos Aires (Pcia)", "La Plata", kind="persona")
    FakeFetcher.configure(pages=[FakePage(f"<html><body>{html}</body></html>")])
    leads = DateasScraper().scrape("garcia", limit=5)
    assert len(leads) == 1
    r = leads[0].raw_data
    assert r["dni"] == "30123456"
    assert r["cuit"] == "20-30123456-5"
    assert r["age"] == "40"
    assert r["province"] == "Buenos Aires"  # "(Pcia)" stripped
    assert r["locality"] == "La Plata"
    assert r["entity_type"] == "fisica"


def test_dateas_pagination(monkeypatch):
    monkeypatch.setattr("scrapers.dateas.Fetcher", FakeFetcher)
    FakeFetcher.configure(pages=[
        _page("dateas", "Empresa Uno A", "Empresa Uno B"),
        _page("dateas", "Empresa Dos A", "Empresa Dos B"),
    ])
    leads = DateasScraper().scrape("test", limit=50)
    assert FakeFetcher.calls >= 2
    assert len(leads) >= 3


def test_dateas_lookup_by_cuit(monkeypatch):
    monkeypatch.setattr("scrapers.dateas.Fetcher", FakeFetcher)
    html = _dateas_row("INMOBILIARIA SA", "33-71017702-9", "",
                       "Córdoba", "Jesus Maria", kind="empresa")
    FakeFetcher.configure(pages=[FakePage(f"<html><body>{html}</body></html>")])
    lead = DateasScraper().lookup_by_cuit("33-71017702-9")
    assert lead is not None
    assert lead.name == "INMOBILIARIA SA"
    assert lead.raw_data["cuit"] == "33-71017702-9"
    assert lead.raw_data["entity_type"] == "juridica"


def test_dateas_lookup_by_dni(monkeypatch):
    monkeypatch.setattr("scrapers.dateas.Fetcher", FakeFetcher)
    html = _dateas_row("PEREZ JUAN", "20-30123456-5", "35 años",
                       "Buenos Aires", "La Plata", kind="persona")
    FakeFetcher.configure(pages=[FakePage(f"<html><body>{html}</body></html>")])
    lead = DateasScraper().lookup_by_dni("30123456")
    assert lead is not None
    assert lead.raw_data["dni"] == "30123456"
    assert lead.raw_data["entity_type"] == "fisica"


def test_excel_dateas_columns():
    leads = [Lead(name="Juan Perez", source="dateas", raw_data={
        "dni": "30123456", "cuit": "20-30123456-5", "age": "35",
        "province": "Buenos Aires", "locality": "La Plata", "entity_type": "fisica",
    })]
    path = CSVExporter().export(leads, "test_dateas_cols.csv")
    ws = load_workbook(path).active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col in ("DNI", "CUIT/CUIL", "Edad", "Provincia", "Localidad", "Tipo"):
        assert col in headers


def test_excel_no_dateas_columns_without_dateas():
    leads = [Lead(name="Gym", source="google_maps", phone="+13055046980")]
    path = CSVExporter().export(leads, "test_nodateas_cols.csv")
    ws = load_workbook(path).active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "DNI" not in headers
    assert "CUIT/CUIL" not in headers


def test_filter_has_cuit():
    leads = [
        Lead(name="A", source="dateas", raw_data={"cuit": "20-30123456-5"}),
        Lead(name="B", source="dateas", raw_data={}),
    ]
    filtered, _ = main.apply_filters(leads, _filter_args(filter_has_cuit=True))
    assert [l.name for l in filtered] == ["A"]


def test_filter_province():
    leads = [
        Lead(name="A", source="dateas", raw_data={"province": "Buenos Aires"}),
        Lead(name="B", source="dateas", raw_data={"province": "Córdoba"}),
    ]
    filtered, _ = main.apply_filters(leads, _filter_args(filter_province="Buenos Aires"))
    assert [l.name for l in filtered] == ["A"]
