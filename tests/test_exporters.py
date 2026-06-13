import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from config.settings import settings
from exporters.csv_exporter import CSVExporter
from exporters.json_exporter import JSONExporter


@pytest.fixture(autouse=True)
def output_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "OUTPUT_DIR", str(tmp_path))
    return tmp_path


def test_xlsx_export_roundtrip(sample_leads):
    path = Path(CSVExporter().export(sample_leads, "test.csv"))
    assert path.suffix == ".xlsx"
    assert path.exists()

    sheet = load_workbook(path).active
    assert sheet.title == "ScrapBro Leads"
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Name", "Email", "Phone", "Website", "Address", "Category", "Rating", "Source"]
    assert sheet.cell(row=2, column=1).value == "Gym A"
    assert sheet.cell(row=2, column=3).value == "+13055046980"


def test_xlsx_header_formatting(sample_leads):
    path = Path(CSVExporter().export(sample_leads, "fmt.csv"))
    sheet = load_workbook(path).active
    header_cell = sheet.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb.endswith("1F3864")
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:H1"


def test_xlsx_totals_row(sample_leads):
    path = Path(CSVExporter().export(sample_leads, "totals.csv"))
    sheet = load_workbook(path).active
    totals_row = len(sample_leads) + 2
    value = sheet.cell(row=totals_row, column=1).value
    assert value.startswith(f"Total: {len(sample_leads)} leads")
    assert "Con email:" in value
    assert "Con teléfono:" in value


def test_xlsx_export_empty_list():
    path = Path(CSVExporter().export([], "empty.csv"))
    assert path.exists()
    sheet = load_workbook(path).active
    assert sheet.cell(row=2, column=1).value.startswith("Total: 0 leads")


def test_json_export_roundtrip(sample_leads):
    path = Path(JSONExporter().export(sample_leads, "test.json"))
    assert path.exists()
    data = json.loads(path.read_text(encoding="utf-8"))
    assert len(data) == len(sample_leads)
    assert data[0]["name"] == "Gym A"
    assert data[2]["source"] == "instagram"


def test_json_export_empty_list():
    path = Path(JSONExporter().export([], "empty.json"))
    assert json.loads(path.read_text(encoding="utf-8")) == []
